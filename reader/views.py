from django.http import JsonResponse
from django.shortcuts import render, redirect
from collections import defaultdict
import openpyxl
import json
import copy
import reader.data as subject_data

def ge_not_list(request):
	context = request.session['context']
	return JsonResponse(context['GE_not'], safe=False)

def major_req_not_list(request):
	context = request.session['context']
	return JsonResponse(context['Major_req_not'], safe=False)

def major_sub_not_list(request):
	context = request.session['context']
	return JsonResponse(context['Major_sub_not'], safe=False)

def index(request):
	return render(request, 'reader/index.html')

def upload_or_result(request):
	if 'context' in request.session:
		return render(request, 'reader/result.html', request.session['context'])
	return render(request, 'reader/upload.html')

def delete_file(request):
	del request.session['context']
	return redirect('/upload')

def sort_by_grade():
    global subject_did, GE_not, no_sub
    score_for_grade = {'A+': 4.5, 'A0': 4.0, 'B+': 3.5, 'B0':3.0, 
                    'C+': 2.5, 'C0': 2.0, 'D+':1.5, 'D0': 1.0, 'P': 5, 'F': 0}

    did = dict()
    for key in subject_did.keys():
        if key[-4:] == '(재수)':
            continue
        if key[1:] in no_sub.keys():
            continue
        if key in subject_data.GE['all']:
            # F아닌데 재수강x 교양 검사
            try:
                new = subject_data.GE_change[key[1:]]
                try:
                    new2 = subject_data.GE_change[new]
                    did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
                    continue
                except:
                    did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
                    continue
            except:
                did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]
                continue
        else:
            did[key] = [subject_did[key][0], subject_did[key][1], score_for_grade[subject_did[key][2]], subject_did[key][2], key]

    value = list(did.values())
    value = sorted(value, key=lambda x: x[2])

    json_parse = dict()
    for key in value:
        # key[-1] => 과목명
        if key[-3] >= 3:
            break
        json_parse[key[-1]] = {'subject': key[-1], 'score': key[-2], 'category': subject_did[key[-1]][0]}

    json_list = list(json_parse.values())
    return json_list

def GE_did_not():
    # 미이수 과목
	global student, area_did
	s_num = int(student['student_num'][2:4]) # 학번 필요한가? 최신(이름 바뀐) 과목 추천해주면 될듯
	r_dict = copy.deepcopy(subject_data.GE)
	
	if s_num >= 21:
		r_dict['other_cnt'] = [0, 1]
	
	recommend = []
 
	for sub in area_did['교필']:
		if sub[2] != 'F':
			if sub[0] in r_dict['English']:
				r_dict['English'].remove(sub[0])
				r_dict['English_cnt'][0] += 1

			elif sub[0][:-4] in r_dict['English']:
				r_dict['English'].remove(sub[0][:-4])
				r_dict['English_cnt'][0] += 1

			elif sub[0] in r_dict['other']:
				r_dict['other'].remove(sub[0])
				r_dict['other_cnt'][0] += 1
			elif sub[0][:-4] in r_dict['other']:
				r_dict['other'].remove(sub[0][:-4])
				r_dict['other_cnt'][0] += 1
			else:
				try:
					if r_dict[sub[0]]:
						r_dict[sub[0]][0] += 1
				except:
					pass
	for key in r_dict['for_loop'].keys():
		r_dict['for_loop'][key]	= max(r_dict[key][1] - r_dict[key][0], 0) # 이수해야하는 만큼 만족하지 못하면 양수 들어감

	for key in r_dict['one_point']:
		for i in range(r_dict['for_loop'][key]):
			recommend.append(key) # 1점짜리 추천
   
	English = []
	for sub in r_dict['English']:
		English.append(sub) # set 자료형은 인덱싱 불가
	English = English[::-1] # 거꾸로 돌림
  
	other = []
	if '대학생활과진로' in r_dict['other']:
		r_dict['for_loop']['other_cnt'] = 1
	else:
		r_dict['for_loop']['other_cnt'] = 0
	for sub in r_dict['other']:
		other.append(sub) # 상동

	other = other[::-1] # 상동

	for i in range(r_dict['for_loop']['English_cnt']):
		recommend.append(English[i])
	for i in range(r_dict['for_loop']['other_cnt']):
		recommend.append(other[i])
	recommend_GE = []
	for sub in recommend:
		recommend_GE.append(subject_data.GE_list[sub])

	if(len(recommend_GE)==0):
		recommend_GE.append({'subject': '<p class="text-danger">미수강한 교필이 없습니다.</p>', 'score': '', 'category': ''})
	return recommend_GE

def remove_jaesu(text):
    if text.endswith("(재수)"):
        return text[:-4]
    else:
        return text
def re_list():
	global subject_did
	list_key = []
	for key in subject_did.keys():
		if key[-4:] == '(재수)':
			list_key.append(key[:-4])

	return list_key

def Major_sub():
	global student, area_did

	if(student['major']!='산업경영공학과'):
		Major_sub_not = {}
		Major_sub_not["none"] = {'subject': '<p class="text-danger">산업경영공학과만 지원합니다.</p>', 'score': '', 'category': ''}
		return Major_sub_not
	
	Major_sub_not = copy.deepcopy(subject_data.IME_list)
	notF = set()

	for i in area_did['전필']:
		tmp = remove_jaesu(i[0])
		if i[2]!='F':
			notF.add(copy.deepcopy(tmp))
	for i in area_did['전선']:
		tmp = remove_jaesu(i[0])
		if i[2]!='F':
			notF.add(copy.deepcopy(tmp))

	for sub in notF:
		while(sub in subject_data.IME_change):
			sub = subject_data.IME_change[sub]
		if(sub in Major_sub_not):
			del Major_sub_not[sub]

	if(len(Major_sub_not)==0):
		Major_sub_not["none"] = {'subject': '<p class="text-danger">미수강한 전선이 없습니다.</p>', 'score': '', 'category': ''}
	
	return Major_sub_not

def Major_req(Major_sub_not):
	global student
	s_num = int(student['student_num'][0:4])#학번찾음
	
	Major_req_not = {}
	Major_req = copy.deepcopy(subject_data.IME_REQ[s_num])
	if(student['major']!='산업경영공학과'):
		Major_req_not["none"] = {'subject': '<p class="text-danger">산업경영공학과만 지원합니다.</p>', 'score': '', 'category': ''}
		return Major_req_not
	
	for sub in Major_req:
		while(sub in subject_data.IME_change):
			sub = subject_data.IME_change[sub]
		if(sub in Major_sub_not):
			Major_req_not[sub] = Major_sub_not[sub]
			del Major_sub_not[sub]

	if(len(Major_req_not)==0):
		Major_req_not["none"] = {'subject': '<p class="text-danger">미수강한 전필이 없습니다.</p>', 'score': '', 'category': ''}

	return Major_req_not

def area_change(Major_req_did, Major_sub_did):
	global student
	s_num = int(student['student_num'][0:4])#학번찾음
	Major_req = copy.deepcopy(subject_data.IME_REQ[s_num])
	need_change = []
	sub_did = set(item[0] for item in Major_sub_did)
	req_did = set(item[0] for item in Major_req_did)
	
	if(student['major']!='산업경영공학과'):
		return need_change
	
	for did in Major_req:
		if(did in sub_did):
			need_change.append({"before":did+"(전선)", "after":did+"(전필)"})
		if(did in req_did):
			req_did.remove(did)

	for did in req_did:
		need_change.append({"before":did+"(전필)", "after":did+"(전선)"})

	return need_change

def grad_cond():
	global student
	s_num = int(student['student_num'][0:4])
	major = student['major']
	
	return subject_data.IME_grad[s_num]






def upload_file(request):
	global student, area, short_area, score_need_list, score_for_grade, info_category, year, score_need, score_did, subject_did, area_did, semester_grade, semester_subject, GE_not, no_sub
	file = request.FILES['uploaded_file']
	if file:
		if file.name.endswith('xlsx'):
			wb = openpyxl.load_workbook(file)
			sheet = wb.active
		#----------------------------------------------------------------------------------------------
			# already set
			area = {'교필영역', '교선영역', '전필영역', 
					'전선영역', '복필영역', '복전영역', '부전공영역', 
					'일선영역', '교직영역'} #, '졸업사정결과'}
			short_area = {'교필', '교선', '전필', 
					'전선', '복필', '복전', '부전', 
					'일선', '교직', '채플'}
			score_need_list = {'교양요구학점', '전필요구학점', '전선요구학점', '복수전공필수요구학점', 
							'복수전공요구학점', '부전공요구학점'}
			score_for_grade = {'A+': 4.5, 'A0': 4.0, 'B+': 3.5, 'B0':3.0, 
							'C+': 2.5, 'C0': 2.0, 'D+':1.5, 'D0': 1.0}
			info_category = ['major', 'minor', 'student_num', 'grade', 
						'name', 'score_need', 'score_did']
			info_idx = ['A2', 'F2', 'J2', 'M2', 'O2', 'Y2', 'AB2']
			year = ['2015', '2016', '2017', '2018', '2019', '2020', '2021', 
				'2022', '2023', '2024', '2025', '2026']
			semester = {'1학기', '2학기'}
			season_semester = ['여름학기', '겨울학기']

			# set during runtime
			score_need = defaultdict(int)	# 영역별 요구 학점
			score_did = defaultdict(float)	# 영역별 이수 학점
			subject_did = defaultdict(list)	# 수강한 과목 모두 (이건 하지 말까 고민중 semester_subject를 모아둔 느낌) => 필요하다!! # [영역, 학점, 등급]
			no_sub = defaultdict(list)
			subject_didnot = defaultdict(dict) # 수강하지 않은 필수 과목 => set(필수 과목) 해서 discard 하는방식
			area_did = defaultdict(list) # 영역별 [과목명, 이수학점, 등급, 년도, 학기]

			semester_grade = defaultdict(dict) # 학년별, 학기별 평점(grade), 이수 학점(score)
			for i in range(1, 6):	# 학년별
				for j in range(1, 3):	# 학기별
					semester_grade[str(i)+'_'+str(j)] = defaultdict(float)

			semester_subject = defaultdict(dict) # 학년별, 학기별 수강 과목
			for i in range(1, 6):	# 학년별
				for j in range(1, 3):	# 학기별
					semester_subject[str(i)+'_'+str(j)] = list()

			student = dict()
			for i in range(7):
				student[info_category[i]] = sheet[info_idx[i]].value

			ex = set()
			season_ex = set()
			j = 1
			while j < sheet.max_row + 1: # 수강 학기 매칭 ex) 1학년 1학기 => 19년도 1학기
				if sheet[str('X'+str(j))].value in year:
					if sheet[str('Z'+str(j))].value in semester:
						ex.add((str(sheet[str('X'+str(j))].value), str(sheet[str('Z'+str(j))].value)[0]))	# ex) (2019,1)
					elif sheet[str('Z'+str(j))].value in season_semester:
						season_ex.add((str(sheet[str('X'+str(j))].value), str(sheet[str('Z'+str(j))].value)))
				if sheet[str('A'+str(j))].value and sheet[str('A'+str(j))].value[:-3] in score_need_list: 
					score_need[str(sheet[str('A'+str(j))].value[:-3])] = int(str(sheet[str('N'+str(j))].value))
					score_need[str(sheet[str('P'+str(j))].value[:-3])] = int(str(sheet[str('W'+str(j))].value))
				j += 1

		#   if sheet[str('Z'+str(j))].value in semester:
		# 				ex.add((str(sheet[str('X'+str(j))].value), str(sheet[str('Z'+str(j))].value)[0]))	# ex) (2019,1)
		# 			elif sheet[str('Z'+str(j))].value in season_semester:
		# 		j += 1
			student_semester = []
			for i in ex:
				student_semester.append(i)
			student_semester.sort()
		
			student_season_semester = []
			for i in season_ex:
				student_season_semester.append(i)
			student_season_semester.sort()

			semester_dict = dict()
			j = 0
			for i in range(len(student_semester)):
				j += 1
				semester_dict[student_semester[i]] = str((i//2)+1)+'_'+str(j) # ex) [('2019', '1')] = (1, 1) 
				j %= 2
			j = 0
			for i in range(len(year)):
				semester_grade[year[i]+'_'+season_semester[j%2]] = defaultdict(float)
				semester_subject[year[i]+'_'+season_semester[j%2]] = list()
				semester_dict[(year[i], season_semester[j%2])] = year[i]+'_'+season_semester[j%2] # ex) [('2019', '여름학기')] = (2019, '여름학기')
				j += 1

			# semester_season_dict = dict()
			# for i in range(len(student_season_semester)):
			# 	semester_season_dict[student_season_semester[i]] = ((i//2)+1, student_season_semester[i][-1]) # ex) [('2019', '여름학기')] = (2019, '여름학기')

			j = 1
			while j < sheet.max_row + 1: # just for the example

				if sheet[str('A'+str(j))].value and sheet[str('A'+str(j))].value[:3] == '교필:':
					total_data = list(sheet[str('A'+str(j))].value.split(' '))
					for data in total_data:
						sub, score = map(str, data.split(':'))
						score_did[sub] += float(score)

				if sheet[str('A'+str(j))].value in area: # ex) 교필영역이 area집합에 있으면
					j += 2
			# 영역  개설학부(과)  이수구분	강좌명	학점  등급	년도  학기  비고
			# A		B			G		I    S	  U    X	Z	AB
			# F면 학점은 0으로, 탭에는 나오게
			# 재수강하지 않았으면 추천 목록에 띄우기

					while sheet[str('G'+str(j))].value in short_area:	# ex) 해당 영역(교필영역)에 수강한 과목이 있으면
						try:
							ex_semester = (semester_dict[(sheet[str('X'+str(j))].value, sheet[str('Z'+str(j))].value[0])])
						except:
							ex_semester = (str(sheet[str('X'+str(j))].value)+'_'+sheet[str('Z'+str(j))].value)
						if sheet[str('U'+str(j))].value != 'P' == 'W':
							continue
						if sheet[str('U'+str(j))].value != 'P' and sheet[str('U'+str(j))].value != 'F': # F맞으면 학점이 하이폰(-)으로 나오나요?
							try:
								semester_grade[ex_semester]['S'] += int(sheet[str('S'+str(j))].value) # 해당 과목 이수 학점
								semester_grade[ex_semester]['G'] += int(sheet[str('S'+str(j))].value)*score_for_grade[sheet[str('U'+str(j))].value]
							except:
								pass

						if sheet[str('U'+str(j))].value == 'P':
							semester_grade[ex_semester]['P'] += int(sheet[str('S'+str(j))].value)

						# [영역, 학점, 등급]

						subject_did[sheet[str('I'+str(j))].value] = [sheet[str('G'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value]
						if sheet[str('U'+str(j))].value[-4:] == 'F' or sheet[str('S'+str(j))].value == '-':
							no_sub[sheet[str('I'+str(j))].value[1:]] = [sheet[str('G'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value]
						
						semester_subject[ex_semester].append(sheet[str('I'+str(j))].value)
						# [과목명, 이수학점, 등급, 년도, 학기]
						area_did[sheet[str('G'+str(j))].value].append([sheet[str('I'+str(j))].value, sheet[str('S'+str(j))].value, sheet[str('U'+str(j))].value, sheet[str('X'+str(j))].value, sheet[str('Z'+str(j))].value])

						j += 1	# excel 다음 행으로
						# if sheet[str('A'+str(j+2))].value and sheet[str('A'+str(j+2))].value[:-3] in score_need_list: 
						# 	score_need[str(sheet[str('A'+str(j+2))].value[:-3])] = int(str(sheet[str('N'+str(j+2))].value))
						# 	score_need[str(sheet[str('P'+str(j+2))].value[:-3])] = int(str(sheet[str('W'+str(j+2))].value))
						# 	break	# 요구학점 나오면 해당 영역 종료

				else:
					j += 1

			# print(f'score_need \n{score_need}\n')
			# print(f'score_did \n{score_did}\n')	
			# print(f'subject_did \n{subject_did}\n')
			# print(f'student \n{student}\n')
			# print(f'semester_subject \n{semester_subject}\n')
			# print(f'semester_dict \n{semester_dict}\n')

			for i in range(1, 6):
				for j in range(1, 3):
					if semester_grade[str(i)+'_'+str(j)]['S']:
						semester_grade[str(i)+'_'+str(j)]['G'] = round(semester_grade[str(i)+'_'+str(j)]['G'] / semester_grade[str(i)+'_'+str(j)]['S'], 2)
						semester_grade[str(i)+'_'+str(j)]['S'] += semester_grade[(str(i)+'_'+str(j))]['P']  # 계산할때는 패논패 계산 없이 함

			# print(f'semester_grade \n{semester_grade}\n')
			# print(f'area_did \n{area_did}\n')
		
			total_avg = sheet[str('H'+str(sheet.max_row - 1))].value # 총 평점 평균
			church = sheet[str('L'+str(sheet.max_row - 1))].value # 채플
			ratio = dict()
			ratio['전필'] = min(100.0, round(score_need['전필이수학점'] / score_need['전필요구학점'], 2)*100)
			ratio['전선'] = min(100.0, round(score_need['전선이수학점'] / score_need['전선요구학점'], 2)*100)
			ratio['교양'] = min(100.0, round(score_need['교양이수학점'] / score_need['교양요구학점'], 2)*100)
			# 교필은 아직임
			# print(f'ratio: \n{ratio}')

			grade_key = {'A+': 'AP', 'A0': 'A', 'B+': 'BP', 'B0': 'B', 
							'C+': 'CP', 'C0': 'C', 'D+': 'DP', 'D0': 'D', 'F': 'F'}
			
			ratio['등급'] = dict()
			for key in grade_key.values():
				ratio['등급'][key] = 0
		
			re_sub = set()
			cnt = 0
			for sub in subject_did:
				if sub[-4:] == "(재수)":
					re_sub.add(sub)
					re_sub.add(sub[:-4])
					try:
						if subject_did[sub][-1] != 'P':
							ratio['등급'][grade_key[subject_did[sub][-1]]] += 1
							cnt += 1
					except:
						pass

			for sub in subject_did:
				if sub[1:] not in re_sub and sub not in re_sub: # 별과목, 과목 인지 몰라욤
					try:
						if subject_did[sub][-1] != 'P':
							ratio['등급'][grade_key[subject_did[sub][-1]]] += 1
							cnt += 1
					except:
						pass
					
			for key in ratio['등급'].keys():
				ratio['등급'][key] = round(ratio['등급'][key] / cnt * 100, 2)

			major_grade = defaultdict(float)

			for data in area_did['전필']:
				if data[2] != 'P' and data[2] != 'F':
					major_grade['전필'] += float(score_for_grade[data[2]])*float(data[1])
					major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
			
			major_grade['전필'] = round(major_grade['전필'] / score_did['전필'], 2)
			
			for data in area_did['전선']:
				if data[2] != 'P' and data[2] != 'F':
					try:
						major_grade['전선'] += float(score_for_grade[data[2]])*float(data[1])
						major_grade['전공'] += float(score_for_grade[data[2]])*float(data[1])
					except:
						pass
			
			major_grade['전선'] = round(major_grade['전선'] / score_did['전선'], 2)
			major_grade['전공']	= round(major_grade['전공'] / (score_did['전선']+score_did['전필']), 2)
		
			sorted_subject = [] # 성적순으로 정렬된것 만들어야함
		
			GE_not = GE_did_not()
			#sorted_grade = sort_by_grade()
			Major_sub_not = Major_sub()
			Major_req_not = list(Major_req(Major_sub_not).values())
			Major_sub_not = list(Major_sub_not.values())
			need_change = area_change(area_did['전필'], area_did['전선'])
			sorted_grade = sort_by_grade() # 재수강 추천

			context = {'area_did': area_did, 
					'semester_grade': semester_grade,
					'semester_subject': semester_subject,
					'student': student,
					'subject_did': subject_did,
					'score_did': score_did,
					'score_need': score_need,
					'total_avg': total_avg,
					'church': church,
					'ratio': ratio,
					'major_grade': major_grade,
					'sorted_subject': sorted_subject,
					'GE_not': GE_not,
					'Major_sub_not' : Major_sub_not,
					'Major_req_not' : Major_req_not,
					'grad' : grad_cond(),
					'sorted_grade': sorted_grade, # 이수한 과목 중 성적 낮은것부터 리스트로
					'need_change': need_change,
					'no_sub': no_sub,
			}
			request.session["context"] = context

	for i in context:
		print(i," ",context[i])
		print()

	return redirect('/upload')


